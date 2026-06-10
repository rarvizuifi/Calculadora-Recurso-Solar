"""
app.py v2.1 — Servidor Flask para el Motor Solar Fotovoltaico mejorado.

Cambios v2.1 sobre v2.0 de Eugenio:
  - Import añadido: fetch_tambiente_profile
  - api_solar(): lee chi y NOCT en lugar de T_coeff; cachea perfil NASA POWER;
    pasa t_amb_profile a run_solar_engine
  - El resto del archivo es idéntico al v2.0 de Eugenio

Endpoints:
  GET  /                        → index.html
  GET  /api/plant_types         → tipos de planta industrial disponibles
  GET  /api/climate_cities      → ciudades climáticas disponibles
  POST /api/demand              → perfil de demanda (35,040 pts)
  POST /api/solar               → motor Jensen v2.1 (con pérdidas, baterías, cortes)
  POST /api/optimal_angles      → ángulos óptimos de instalación por lat/lon
  GET  /api/download/excel      → Excel con todos los datos (6 hojas)
  GET  /api/glossary            → glosario de parámetros y resultados (JSON)
"""

import io
import re
import math
import datetime
import traceback
import unicodedata
import numpy as np
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from demand_profile import generate_demand_profile, PLANT_PROFILES
from solar_engine import run_solar_engine, get_climate_cities
from solar_engine import fetch_tambiente_profile   # NUEVO v2.1

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

_cache = {}
DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

GLOSSARY = {
    "parametros_entrada": {
        "lat":               {"desc": "Latitud geográfica [°]. Norte positivo. Ej: 25.67 para Monterrey.", "unidad": "°"},
        "lon":               {"desc": "Longitud geográfica [°]. Oeste negativo. Ej: -100.31 para Monterrey.", "unidad": "°"},
        "alt":               {"desc": "Altitud sobre el nivel del mar [m]. Influye en la masa de aire y la irradiancia.", "unidad": "m"},
        "eta":               {"desc": "Ignorado en v2.1. eta_ref se deduce de P_nominal/(1000·A).", "unidad": "fracción 0-1"},
        "area_m2":           {"desc": "Área activa de un panel solar [m²]. Ej: 2.0 m².", "unidad": "m²"},
        "n_panels":          {"desc": "Número de paneles en el arreglo fotovoltaico.", "unidad": "unidades"},
        "tilt":              {"desc": "Ángulo de inclinación del panel respecto a la horizontal [°]. 0 = horizontal, 90 = vertical.", "unidad": "°"},
        "azimuth":           {"desc": "Orientación del panel desde el Norte en sentido horario [°]. 0=N, 90=E, 180=S, 270=O.", "unidad": "°"},
        "p_nominal_w":       {"desc": "Potencia nominal (pico) de un panel en STC [W]. Ej: 400 W.", "unidad": "W"},
        "chi":               {"desc": "[NUEVO v2.1] Coeficiente de temperatura de potencia [%/°C]. Monocristalino ≈ 0.35–0.45.", "unidad": "%/°C"},
        "NOCT":              {"desc": "Temperatura de Operación Nominal de Celda [°C]. Típico: 44–48 °C.", "unidad": "°C"},
        "loss_wiring":       {"desc": "Pérdidas por resistencia del cableado DC y AC. Típico: 0.02–0.03.", "unidad": "fracción 0-1"},
        "loss_inverter":     {"desc": "Pérdidas en el inversor (conversión DC→AC). Típico: 0.02–0.04.", "unidad": "fracción 0-1"},
        "loss_dirt":         {"desc": "Pérdidas por suciedad o polvo sobre los paneles. Típico: 0.02–0.05.", "unidad": "fracción 0-1"},
        "loss_mismatch":     {"desc": "Pérdidas por variación entre paneles del mismo arreglo. Típico: 0.01–0.02.", "unidad": "fracción 0-1"},
        "loss_shading":      {"desc": "Pérdidas por sombras sobre el arreglo. Típico: 0.00–0.10.", "unidad": "fracción 0-1"},
        "wind_speed":        {"desc": "Velocidad media del viento en el sitio [m/s].", "unidad": "m/s"},
        "humidity_pct":      {"desc": "Humedad relativa media del ambiente [%].", "unidad": "%"},
        "battery_kwh":       {"desc": "Capacidad total útil del banco de baterías [kWh]. 0 = sin baterías.", "unidad": "kWh"},
        "battery_eta":       {"desc": "Eficiencia round-trip del banco. Típico: 0.90–0.95 para Li-Ion.", "unidad": "fracción 0-1"},
        "battery_dod":       {"desc": "Profundidad de descarga máxima permitida. Ej: 0.80 = 80 %.", "unidad": "fracción 0-1"},
        "battery_cycles":    {"desc": "Vida útil del banco en ciclos completos. Li-Ion NMC ≈ 3000–6000, LFP ≈ 5000–8000.", "unidad": "ciclos"},
        "battery_cost_kwh":  {"desc": "Costo de adquisición de la batería por kWh instalado [USD/kWh].", "unidad": "USD/kWh"},
        "battery_install_pct": {"desc": "Costo de instalación como fracción del costo de equipo. Típico: 0.20–0.35.", "unidad": "fracción 0-1"},
        "outage_freq_yr":    {"desc": "Número esperado de cortes de energía de la red por año.", "unidad": "cortes/año"},
        "outage_avg_h":      {"desc": "Duración media de cada corte de energía [horas].", "unidad": "horas"},
        "outage_loss_pct":   {"desc": "Porcentaje de la demanda no cubierta durante los cortes.", "unidad": "%"},
        "climate_city":      {"desc": "Clave de ciudad de la base de datos climática interna.", "unidad": "texto"},
        "panel_cost_usd":    {"desc": "Costo de adquisición de un panel solar [USD].", "unidad": "USD/panel"},
        "inverter_cost_usd": {"desc": "Costo del inversor [USD].", "unidad": "USD"},
    },
    "resultados": {
        "energia_anual_kWh":         {"desc": "Energía eléctrica total neta generada en el año [kWh].", "unidad": "kWh/año"},
        "energia_anual_MWh":         {"desc": "Energía eléctrica total neta en MWh.", "unidad": "MWh/año"},
        "p_max_kW":                  {"desc": "Potencia pico real alcanzada por el arreglo durante el año [kW].", "unidad": "kW"},
        "p_nominal_total_kW":        {"desc": "Potencia nominal total del sistema = n_paneles × p_nominal_W / 1000 [kWp].", "unidad": "kWp"},
        "factor_capacidad_pct":      {"desc": "Factor de capacidad = E_anual / (P_nominal × 8760 h) × 100.", "unidad": "%"},
        "horas_pico_sol_equiv":      {"desc": "Horas Pico Solar Equivalente = E_anual / P_nominal.", "unidad": "h/año"},
        "irrad_horizontal_kWh_m2":   {"desc": "Irradiación total anual en plano horizontal [kWh/m²·año].", "unidad": "kWh/m²·año"},
        "irrad_poa_kWh_m2":          {"desc": "Irradiación total anual en el Plano del Arreglo (POA) [kWh/m²·año].", "unidad": "kWh/m²·año"},
        "T_cell_media_C":            {"desc": "Temperatura media de celda durante horas de sol [°C].", "unidad": "°C"},
        "T_cell_max_C":              {"desc": "Temperatura máxima de celda alcanzada en el año [°C].", "unidad": "°C"},
        "perdida_termica_pct":       {"desc": "Pérdida porcentual anual de energía debida al efecto de temperatura.", "unidad": "%"},
        "perdida_sistema_pct":       {"desc": "Pérdida total del sistema (cableado + inversor + suciedad + mismatch + sombras) [%].", "unidad": "%"},
        "performance_ratio_pct":     {"desc": "Performance Ratio (PR) del sistema. Típico: 75–85 %.", "unidad": "%"},
        "battery_energy_stored_kWh": {"desc": "Energía total cargada en el banco de baterías durante el año [kWh].", "unidad": "kWh/año"},
        "battery_energy_served_kWh": {"desc": "Energía total descargada / entregada por las baterías al año [kWh].", "unidad": "kWh/año"},
        "battery_cycles_used":       {"desc": "Ciclos equivalentes de carga/descarga usados en el año.", "unidad": "ciclos/año"},
        "battery_life_years":        {"desc": "Vida útil estimada del banco de baterías [años].", "unidad": "años"},
        "battery_capex_usd":         {"desc": "CAPEX total del banco de baterías (equipo + instalación) [USD].", "unidad": "USD"},
        "outage_hours_yr":           {"desc": "Horas totales de corte de energía esperadas por año.", "unidad": "h/año"},
        "outage_energy_lost_kWh":    {"desc": "Energía estimada no servida por cortes de energía al año [kWh].", "unidad": "kWh/año"},
        "total_capex_usd":           {"desc": "CAPEX total del sistema: paneles + inversor + baterías [USD].", "unidad": "USD"},
        "eta_ref":                   {"desc": "[NUEVO v2.1] Eficiencia real deducida del datasheet: P_nominal/(1000·A).", "unidad": "fracción 0-1"},
        "chi_pct_por_c":             {"desc": "[NUEVO v2.1] Coeficiente de temperatura de potencia usado.", "unidad": "%/°C"},
        "fuente_t_amb":              {"desc": "[NUEVO v2.1] Fuente del perfil de T_amb: NASA POWER o estimación sinusoidal.", "unidad": "texto"},
    }
}


# ─── Frontend ─────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


# ─── Glosario ─────────────────────────────────────────────────────────────────
@app.route('/api/glossary', methods=['GET'])
def api_glossary():
    return jsonify({'ok': True, 'glossary': GLOSSARY})


# ─── Ángulos Óptimos ──────────────────────────────────────────────────────────
_DEG = math.pi / 180
_MONTHS_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
_MONTH_DAYS = [17, 47, 75, 105, 135, 162, 198, 228, 258, 288, 318, 344]

def _declination_rad(n: int) -> float:
    return 23.45 * _DEG * math.sin(2 * math.pi * (n - 81) / 365)

def _solar_noon_elevation_rad(lat_deg: float, delta_rad: float) -> float:
    phi = lat_deg * _DEG
    sin_a = math.sin(phi)*math.sin(delta_rad) + math.cos(phi)*math.cos(delta_rad)
    return math.asin(max(-1.0, min(1.0, sin_a)))

def _incidence_angle_rad(lat_deg: float, delta_rad: float, tilt_deg: float) -> float:
    phi  = lat_deg * _DEG
    beta = tilt_deg * _DEG
    cos_t = (math.sin(delta_rad) * math.sin(phi - beta) +
             math.cos(delta_rad) * math.cos(phi - beta))
    return math.acos(max(-1.0, min(1.0, cos_t)))

def _optimal_tilt(lat_deg: float) -> float:
    best_tilt, best_sum = 0.0, -1.0
    for t_half in range(0, 181):
        t = t_half * 0.5
        s = 0.0
        for n in _MONTH_DAYS:
            delta = _declination_rad(n)
            alpha = _solar_noon_elevation_rad(lat_deg, delta)
            if alpha <= 0:
                continue
            theta = _incidence_angle_rad(lat_deg, delta, t)
            ct = math.cos(theta)
            if ct > 0:
                s += ct
        if s > best_sum:
            best_sum, best_tilt = s, t
    return round(best_tilt * 2) / 2

def _estimate_annual_irrad(lat_deg: float, tilt_deg: float) -> float:
    GSC = 1367.0
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    total = 0.0
    phi  = lat_deg * _DEG
    beta = tilt_deg * _DEG
    for m, n in enumerate(_MONTH_DAYS):
        delta = _declination_rad(n)
        Eo    = 1.0 + 0.033 * math.cos(2 * math.pi * n / 365)
        cos_ws = -math.tan(phi) * math.tan(delta)
        if cos_ws >= 1.0:
            continue
        omega_s = math.acos(max(-1.0, min(1.0, cos_ws)))
        N_steps = 48
        daily = 0.0
        for i in range(N_steps + 1):
            omega = -omega_s + (i / N_steps) * 2 * omega_s
            sin_a = (math.sin(phi)*math.sin(delta) +
                     math.cos(phi)*math.cos(delta)*math.cos(omega))
            if sin_a <= 0:
                continue
            alpha = math.asin(max(-1.0, min(1.0, sin_a)))
            AM    = 1.0 / math.sin(alpha)
            tau_b = 0.7 ** (AM ** 0.678)
            G0    = GSC * Eo * math.sin(alpha)
            Gb_h  = max(0.0, G0 * tau_b)
            Gd_h  = max(0.0, G0 * (1 - tau_b) * 0.5)
            cos_t = (math.sin(delta)*math.sin(phi - beta) +
                     math.cos(delta)*math.cos(phi - beta)*math.cos(omega))
            Rb    = (cos_t / math.sin(alpha)
                     if (cos_t > 0 and math.sin(alpha) > 0.01) else 0.0)
            Gb_poa = max(0.0, Gb_h * Rb)
            Gd_poa = Gd_h * (1 + math.cos(beta)) / 2.0
            Gr_poa = (Gb_h + Gd_h) * 0.20 * (1 - math.cos(beta)) / 2.0
            daily += max(0.0, Gb_poa + Gd_poa + Gr_poa)
        daylight_h = omega_s / math.pi * 24.0
        daily_avg  = daily / (N_steps + 1)
        total += daily_avg * daylight_h * days_in_month[m] / 1000.0
    return round(total, 1)


@app.route('/api/optimal_angles', methods=['POST'])
def api_optimal_angles():
    try:
        d   = request.get_json(force=True)
        lat = float(d.get('lat', 25.67))
        lon = float(d.get('lon', -100.31))
        lat = float(np.clip(lat, -90, 90))

        tilt_opt     = _optimal_tilt(lat)
        azimuth_opt  = 180.0 if lat >= 0 else 0.0
        hemisferio   = 'Norte' if lat >= 0 else 'Sur'

        delta_solst = _declination_rad(172)
        delta_equin = _declination_rad(80)
        elev_solst  = _solar_noon_elevation_rad(lat, delta_solst) / _DEG
        elev_equin  = _solar_noon_elevation_rad(lat, delta_equin) / _DEG
        zenith_solst = 90.0 - elev_solst

        irrad_est   = _estimate_annual_irrad(lat, tilt_opt)

        monthly = []
        for m, n in enumerate(_MONTH_DAYS):
            delta = _declination_rad(n)
            alpha = _solar_noon_elevation_rad(lat, delta) / _DEG
            theta = _incidence_angle_rad(lat, delta, tilt_opt) / _DEG
            cos_t = max(0.0, math.cos(theta * _DEG))
            monthly.append({
                'mes'              : _MONTHS_ES[m],
                'dia_repr'         : n,
                'declinacion_deg'  : round(delta / _DEG, 2),
                'elevacion_solar_deg': round(alpha, 2),
                'angulo_incidencia_deg': round(theta, 2),
                'cos_theta'        : round(cos_t, 4),
            })

        result = {
            'lat'                    : lat,
            'lon'                    : lon,
            'hemisferio'             : hemisferio,
            'tilt_optimo_deg'        : tilt_opt,
            'azimuth_optimo_deg'     : azimuth_opt,
            'azimuth_optimo_desc'    : f"{azimuth_opt:.0f}° — {'Sur (óptimo hemisferio Norte)' if lat >= 0 else 'Norte (óptimo hemisferio Sur)'}",
            'elevacion_solsticio_deg': round(elev_solst, 2),
            'elevacion_equinoccio_deg': round(elev_equin, 2),
            'angulo_cenital_solsticio_deg': round(zenith_solst, 2),
            'irradiacion_poa_anual_kWh_m2': irrad_est,
            'regla_practica'         : f"Tilt ≈ {abs(lat):.1f}° (latitud) ajustado a {tilt_opt:.1f}° para máxima captación anual",
            'monthly_table'          : monthly,
        }

        _cache['optimal_angles'] = result
        return jsonify({'ok': True, **result})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 400


# ─── Tipos de planta ──────────────────────────────────────────────────────────
@app.route('/api/plant_types', methods=['GET'])
def api_plant_types():
    return jsonify({
        'ok': True,
        'types': [{'key': k, 'name': v['name'], 'desc': v['desc']}
                  for k, v in PLANT_PROFILES.items()]
    })


# ─── Ciudades climáticas ──────────────────────────────────────────────────────
@app.route('/api/climate_cities', methods=['GET'])
def api_climate_cities():
    return jsonify({'ok': True, 'cities': get_climate_cities()})


# ─── Perfil de Demanda ────────────────────────────────────────────────────────
@app.route('/api/demand', methods=['POST'])
def api_demand():
    try:
        d            = request.get_json(force=True)
        Pmax         = max(1.0, float(d.get('pmax_kW', 50)))
        FC           = float(np.clip(float(d.get('fc_planta', 0.60)), 0.40, 0.90))
        FP           = float(np.clip(float(d.get('fp_potencia', 0.85)), 0.60, 1.00))
        n_shifts     = int(np.clip(int(d.get('n_shifts', 2)), 1, 3))
        plant_type   = str(d.get('plant_type', 'manufactura_ligera'))
        weekend_f    = float(np.clip(float(d.get('weekend_op_factor', 0.50)), 0.0, 1.0))
        summer_boost = float(np.clip(float(d.get('summer_boost', 1.10)), 1.0, 1.50))

        result = generate_demand_profile(
            Pmax_kW=Pmax, FC_planta=FC, FP_potencia=FP,
            n_shifts=n_shifts, plant_type=plant_type,
            weekend_op_factor=weekend_f, summer_boost=summer_boost,
        )
        _cache['demand'] = result

        return jsonify({
            'ok': True,
            'monthly_avg'  : result['monthly_avg'],
            'monthly_max'  : result['monthly_max'],
            'monthly_min'  : result['monthly_min'],
            'monthly_kWh'  : result['monthly_kWh'],
            'daily_profile': result['daily_weekday'],
            'daily_weekend': result['daily_weekend'],
            'stats'        : result['stats'],
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 400


# ─── Motor Solar v2.1 ─────────────────────────────────────────────────────────
@app.route('/api/solar', methods=['POST'])
def api_solar():
    try:
        d = request.get_json(force=True)

        # Parámetros básicos (sin cambios respecto a v2.0)
        lat         = float(np.clip(float(d.get('lat', 25.67)), -90, 90))
        lon         = float(d.get('lon', -100.31))
        alt         = float(d.get('alt', 538))
        eta_stc     = d.get('eta_stc', None)
        if eta_stc is not None:
            eta_stc = float(np.clip(float(eta_stc), 0.05, 0.50))
        area_m2     = float(d.get('area_m2', 2.0))
        n_panels    = max(1, int(d.get('n_panels', 50)))
        tilt        = float(np.clip(float(d.get('tilt', 25.0)), 0, 90))
        azimuth     = float(d.get('azimuth', 180.0))
        p_nominal_w = float(d.get('p_nominal_w', 400))

        # Pérdidas del sistema (sin cambios)
        loss_wiring   = float(np.clip(float(d.get('loss_wiring',   0.02)), 0, 0.20))
        loss_inverter = float(np.clip(float(d.get('loss_inverter', 0.03)), 0, 0.15))
        loss_dirt     = float(np.clip(float(d.get('loss_dirt',     0.03)), 0, 0.20))
        loss_mismatch = float(np.clip(float(d.get('loss_mismatch', 0.01)), 0, 0.10))
        loss_shading  = float(np.clip(float(d.get('loss_shading',  0.00)), 0, 0.50))

        # Modelo térmico — CAMBIO v2.1: chi reemplaza T_coeff; NOCT sin cambios
        NOCT       = float(d.get('NOCT', 45.0))
        chi        = float(np.clip(float(d.get('chi', 0.40)), 0.20, 0.60))  # NUEVO v2.1
        wind_speed = float(d.get('wind_speed', 3.0))
        humidity   = float(d.get('humidity_pct', 55.0))

        # Baterías (sin cambios)
        battery_kwh         = float(d.get('battery_kwh', 0.0))
        battery_eta         = float(np.clip(float(d.get('battery_eta', 0.92)), 0.70, 0.99))
        battery_dod         = float(np.clip(float(d.get('battery_dod', 0.80)), 0.20, 1.00))
        battery_cycles      = int(d.get('battery_cycles', 4000))
        battery_cost_kwh    = float(d.get('battery_cost_kwh', 300.0))
        battery_install_pct = float(d.get('battery_install_pct', 0.25))

        # Cortes de energía (sin cambios)
        outage_freq    = float(d.get('outage_freq_yr', 0.0))
        outage_avg_h   = float(d.get('outage_avg_h', 2.0))
        outage_loss_pct= float(d.get('outage_loss_pct', 80.0))

        # Clima (sin cambios)
        climate_city = d.get('climate_city', None)

        # Costos (sin cambios)
        panel_cost    = float(d.get('panel_cost_usd', 0.0))
        inverter_cost = float(d.get('inverter_cost_usd', 0.0))

        # NUEVO v2.1: T_amb desde NASA POWER, cacheado por coordenada
        cache_key = f"tambiente_{round(lat, 2)}_{round(lon, 2)}"
        if cache_key not in _cache:
            _cache[cache_key] = fetch_tambiente_profile(lat, lon)

        result = run_solar_engine(
            lat=lat, lon=lon, alt=alt,
            eta_stc=eta_stc, area_m2=area_m2, n_panels=n_panels, tilt=tilt, azimuth=azimuth,
            p_nominal_w=p_nominal_w,
            loss_wiring=loss_wiring, loss_inverter=loss_inverter,
            loss_dirt=loss_dirt, loss_mismatch=loss_mismatch,
            loss_shading=loss_shading,
            NOCT=NOCT, chi=chi,                        # NUEVO v2.1
            wind_speed=wind_speed, humidity_pct=humidity,
            t_amb_profile=_cache[cache_key],            # NUEVO v2.1
            battery_kwh=battery_kwh, battery_eta=battery_eta,
            battery_dod=battery_dod, battery_cycles=battery_cycles,
            battery_cost_kwh=battery_cost_kwh,
            battery_install_pct=battery_install_pct,
            outage_freq_yr=outage_freq, outage_avg_h=outage_avg_h,
            outage_loss_pct=outage_loss_pct,
            climate_city=climate_city,
            panel_cost_usd=panel_cost, inverter_cost_usd=inverter_cost,
        )

        # Balance con demanda (sin cambios respecto a v2.0)
        balance = None
        if 'demand' in _cache:
            dem_arr = np.array(_cache['demand']['demand_kW'])
            gen_arr = np.array(result['P_kw_arr'])
            exceso  = np.maximum(gen_arr - dem_arr, 0)
            deficit = np.maximum(dem_arr - gen_arr, 0)
            e_dem   = float(np.sum(dem_arr) * 0.25)
            e_gen   = float(np.sum(gen_arr) * 0.25)
            cob     = min(e_gen / e_dem * 100, 100) if e_dem > 0 else 0

            monthly_balance, monthly_cob = [], []
            idx = 0
            for nd in DAYS_IN_MONTH:
                np_ = nd * 96
                eg  = float(np.sum(gen_arr[idx:idx + np_]) * 0.25)
                ed  = float(np.sum(dem_arr[idx:idx + np_]) * 0.25)
                monthly_balance.append(round(eg - ed, 2))
                monthly_cob.append(round(min(eg / ed * 100, 100) if ed > 0 else 0, 2))
                idx += np_

            balance = {
                'energia_demanda_kWh' : round(e_dem, 2),
                'energia_generada_kWh': round(e_gen, 2),
                'cobertura_pct'       : round(cob, 2),
                'exceso_kWh'          : round(float(np.sum(exceso) * 0.25), 2),
                'deficit_kWh'         : round(float(np.sum(deficit) * 0.25), 2),
                'monthly_balance'     : monthly_balance,
                'monthly_cobertura'   : monthly_cob,
            }

        _cache['solar'] = result

        return jsonify({
            'ok'                 : True,
            'monthly_gtot_avg'   : result['monthly_gtot_avg'],
            'monthly_gtot_max'   : result['monthly_gtot_max'],
            'monthly_gen_kWh'    : result['monthly_gen_kWh'],
            'monthly_bat_charge' : result['monthly_bat_charge'],
            'monthly_bat_discharge': result['monthly_bat_discharge'],
            'daily_gtot_summer'  : result['daily_gtot_summer'],
            'daily_p_summer'     : result['daily_p_summer'],
            'stats'              : result['stats'],
            'balance'            : balance,
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 400


# ─── Descarga Excel (idéntico a v2.0 de Eugenio) ─────────────────────────────
@app.route('/api/download/excel', methods=['GET'])
def api_download_excel():
    if 'solar' not in _cache:
        return jsonify({'ok': False, 'error': 'Ejecuta primero el Motor Solar.'}), 400

    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    solar  = _cache['solar']
    demand = _cache.get('demand')

    hours  = solar['hours']
    Gtot   = solar['Gtot_arr']
    P_kw   = solar['P_kw_arr']
    P_gross= solar['P_kw_gross_arr']
    T_cell = solar['T_cell_arr']
    dem_kw = demand['demand_kW'] if demand else [0] * len(hours)

    base_dt = datetime.datetime(2024, 1, 1, 0, 0)
    fechas  = [(base_dt + datetime.timedelta(hours=h)).strftime('%Y-%m-%d %H:%M')
               for h in hours]

    wb = openpyxl.Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="0D1526")
    COL_FILLS  = {
        'fecha'  : PatternFill("solid", fgColor="111827"),
        'irrad'  : PatternFill("solid", fgColor="1a1f0a"),
        'gen'    : PatternFill("solid", fgColor="0a1a14"),
        'dem'    : PatternFill("solid", fgColor="1a100a"),
        'bal'    : PatternFill("solid", fgColor="0a0a1a"),
        'temp'   : PatternFill("solid", fgColor="1a0a0a"),
    }
    HDR_FONT   = Font(name='Calibri', bold=True, color='F97316', size=10)
    DATA_FONT  = Font(name='Calibri', size=9)
    TITLE_FONT = Font(name='Calibri', bold=True, color='FBBF24', size=13)
    thin       = Side(style='thin', color='1E293B')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = HDR_FONT
        c.fill  = fill or HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        return c

    def data_cell(ws, row, col, val, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = DATA_FONT
        c.border = border
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    # ── Hoja 1: Parámetros ────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = '1. Parámetros'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 24

    s = solar['stats']
    title = ws1.cell(row=1, column=1, value='Motor Solar FV v2.1 — Parámetros')
    title.font = TITLE_FONT
    ws1.merge_cells('A1:B1')

    params = [
        ('── SISTEMA FV ──', ''),
        ('Latitud',                  f"{s['lat']} °"),
        ('Longitud',                 f"{s['lon']} °"),
        ('Altitud',                  f"{s['alt']} m s.n.m."),
        ('Ciudad climática',         s.get('climate_city', 'manual')),
        ('η_ref (deducido del datasheet)', f"{s.get('eta_ref', s.get('eta', 0))*100:.2f} %"),
        ('Área por panel',           f"{s['area_m2']} m²"),
        ('Potencia nominal / panel', f"{s['potencia_nominal_W_panel']} W"),
        ('Número de paneles',        s['n_paneles']),
        ('Inclinación (tilt)',       f"{s['tilt']} °"),
        ('Azimut',                   f"{s['azimuth']} °"),
        ('', ''),
        ('── MODELO TÉRMICO v2.1 ──', ''),
        ('NOCT',                     f"{s.get('NOCT', 45.0)} °C"),
        ('Coef. temperatura χ',      f"{s.get('chi_pct_por_c', 0.40)} %/°C"),
        ('Velocidad de viento',      f"{s.get('wind_speed_ms', 3.0)} m/s"),
        ('Humedad relativa',         f"{s.get('humidity_pct', 55.0)} %"),
        ('T_amb media',              f"{s.get('T_amb_avg_C', 20.0)} °C"),
        ('Fuente T_amb',             s.get('fuente_t_amb', '—')),
        ('', ''),
        ('── PÉRDIDAS DEL SISTEMA ──', ''),
        ('Pérd. cableado',           f"{s.get('loss_wiring_pct', 2.0)} %"),
        ('Pérd. inversor',           f"{s.get('loss_inverter_pct', 3.0)} %"),
        ('Pérd. suciedad',           f"{s.get('loss_dirt_pct', 3.0)} %"),
        ('Pérd. mismatch',           f"{s.get('loss_mismatch_pct', 1.0)} %"),
        ('Pérd. sombras',            f"{s.get('loss_shading_pct', 0.0)} %"),
        ('Performance Ratio (PR)',   f"{s.get('performance_ratio_pct', 91.0):.1f} %"),
        ('', ''),
        ('── BATERÍAS ──', ''),
        ('Capacidad del banco',      f"{s.get('battery_kwh', 0.0):.1f} kWh"),
        ('Eficiencia round-trip',    f"{s.get('battery_eta_rt', 0.92)*100:.0f} %"),
        ('CAPEX baterías',           f"USD {s.get('battery_capex_usd', 0.0):,.0f}"),
        ('Vida útil estimada',       f"{s.get('battery_life_years', 0.0):.1f} años"),
        ('', ''),
        ('── CORTES DE ENERGÍA ──', ''),
        ('Frecuencia de cortes',     f"{s.get('outage_freq_yr', 0.0):.1f} /año"),
        ('Duración media de corte',  f"{s.get('outage_avg_h', 0.0):.1f} h"),
        ('Horas de corte / año',     f"{s.get('outage_hours_yr', 0.0):.2f} h"),
        ('', ''),
        ('── CAPEX TOTAL ──', ''),
        ('CAPEX paneles',            f"USD {s.get('panel_cost_usd', 0.0):,.0f}"),
        ('CAPEX inversor',           f"USD {s.get('inverter_cost_usd', 0.0):,.0f}"),
        ('CAPEX baterías',           f"USD {s.get('battery_capex_usd', 0.0):,.0f}"),
        ('CAPEX TOTAL',              f"USD {s.get('total_capex_usd', 0.0):,.0f}"),
    ]
    if demand:
        ds = demand['stats']
        params += [
            ('', ''),
            ('── PLANTA INDUSTRIAL ──', ''),
            ('Tipo de planta',       ds.get('plant_name', '—')),
            ('Demanda máxima',       f"{ds['pmax_kW']:.1f} kW"),
            ('Turnos de operación',  ds.get('n_shifts', '—')),
            ('Factor de carga',      f"{ds['FC_planta']*100:.0f} %"),
            ('Factor de potencia',   f"{ds['FP_potencia']:.2f}"),
        ]

    for i, (k, v) in enumerate(params, start=3):
        ck = ws1.cell(row=i, column=1, value=k)
        cv = ws1.cell(row=i, column=2, value=v)
        if k.startswith('──'):
            ck.font = Font(name='Calibri', bold=True, color='F97316', size=10); cv.value = ''
        else:
            ck.font = DATA_FONT; cv.font = Font(name='Calibri', bold=True, color='F1F5F9', size=9)
        for c in [ck, cv]:
            c.border = border; c.alignment = Alignment(vertical='center')

    # ── Hoja 2: KPIs ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('2. KPIs')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 14
    t2 = ws2.cell(row=1, column=1, value='Resumen Ejecutivo — KPIs del Sistema v2.1')
    t2.font = TITLE_FONT; ws2.merge_cells('A1:C1')

    kpis = [
        ('☀️  GENERACIÓN SOLAR', '', ''),
        ('Energía anual generada', f"{s['energia_anual_kWh']:,.0f}", 'kWh/año'),
        ('Energía anual generada', f"{s['energia_anual_MWh']:,.2f}", 'MWh/año'),
        ('Potencia pico del sistema', f"{s['p_nominal_total_kW']:,.2f}", 'kWp'),
        ('Potencia máxima generada', f"{s['p_max_kW']:,.2f}", 'kW'),
        ('Factor de capacidad', f"{s['factor_capacidad_pct']:,.2f}", '%'),
        ('Horas pico solar equiv.', f"{s['horas_pico_sol_equiv']:,.0f}", 'h/año'),
        ('Horas con generación > 0', f"{s['n_horas_generacion']:,.0f}", 'h/año'),
        ('Irradiación horizontal', f"{s['irrad_horizontal_kWh_m2']:,.0f}", 'kWh/m²·año'),
        ('Irradiación POA', f"{s['irrad_poa_kWh_m2']:,.0f}", 'kWh/m²·año'),
        ('', '', ''),
        ('🌡️  TEMPERATURA Y PÉRDIDAS (v2.1)', '', ''),
        ('η_ref deducido del datasheet', f"{s.get('eta_ref', 0)*100:.2f}", '%'),
        ('Coef. temp. χ utilizado', f"{s.get('chi_pct_por_c', 0.40)}", '%/°C'),
        ('Fuente T_amb', s.get('fuente_t_amb', '—'), '—'),
        ('Temp. celda media (horas sol)', f"{s.get('T_cell_media_C',0):.1f}", '°C'),
        ('Temp. celda máxima', f"{s.get('T_cell_max_C',0):.1f}", '°C'),
        ('Pérdida térmica anual', f"{s.get('perdida_termica_pct',0):.2f}", '%'),
        ('Pérdida de sistema (PR)', f"{s.get('perdida_sistema_pct',0):.2f}", '%'),
        ('Performance Ratio', f"{s.get('performance_ratio_pct',0):.2f}", '%'),
        ('', '', ''),
        ('🔋  BATERÍAS', '', ''),
        ('Capacidad del banco', f"{s.get('battery_kwh',0):.1f}", 'kWh'),
        ('Energía almacenada / año', f"{s.get('battery_energy_stored_kWh',0):,.0f}", 'kWh/año'),
        ('Energía servida / año', f"{s.get('battery_energy_served_kWh',0):,.0f}", 'kWh/año'),
        ('Ciclos usados / año', f"{s.get('battery_cycles_used',0):.1f}", 'ciclos/año'),
        ('Vida útil estimada', f"{s.get('battery_life_years',0):.1f}", 'años'),
        ('CAPEX baterías', f"USD {s.get('battery_capex_usd',0):,.0f}", 'USD'),
        ('', '', ''),
        ('⚡  CORTES DE ENERGÍA', '', ''),
        ('Horas de corte / año', f"{s.get('outage_hours_yr',0):.2f}", 'h/año'),
        ('Energía no servida', f"{s.get('outage_energy_lost_kWh',0):,.0f}", 'kWh/año'),
        ('', '', ''),
        ('💰  CAPEX DEL SISTEMA', '', ''),
        ('CAPEX paneles', f"USD {s.get('panel_cost_usd',0):,.0f}", 'USD'),
        ('CAPEX inversor', f"USD {s.get('inverter_cost_usd',0):,.0f}", 'USD'),
        ('CAPEX baterías', f"USD {s.get('battery_capex_usd',0):,.0f}", 'USD'),
        ('CAPEX TOTAL', f"USD {s.get('total_capex_usd',0):,.0f}", 'USD'),
    ]

    if demand:
        ds = demand['stats']
        dem_arr = np.array(demand['demand_kW'])
        gen_arr = np.array(solar['P_kw_arr'])
        e_dem   = float(np.sum(dem_arr) * 0.25)
        e_gen   = float(np.sum(gen_arr) * 0.25)
        cob     = min(e_gen / e_dem * 100, 100) if e_dem > 0 else 0
        exceso  = float(np.sum(np.maximum(gen_arr - dem_arr, 0)) * 0.25)
        deficit = float(np.sum(np.maximum(dem_arr - gen_arr, 0)) * 0.25)
        kpis += [
            ('', '', ''),
            ('🏭  DEMANDA Y BALANCE', '', ''),
            ('Energía demandada', f"{e_dem:,.0f}", 'kWh/año'),
            ('Cobertura solar', f"{cob:.1f}", '%'),
            ('Excedente solar', f"{exceso:,.0f}", 'kWh/año'),
            ('Déficit (red)', f"{deficit:,.0f}", 'kWh/año'),
        ]

    hdr(ws2, 2, 1, 'Indicador'); hdr(ws2, 2, 2, 'Valor'); hdr(ws2, 2, 3, 'Unidad')
    for i, (k, v, u) in enumerate(kpis, start=3):
        if any(k.endswith(x) for x in ['SOLAR','TEMPERATURA Y PÉRDIDAS (v2.1)','BATERÍAS',
                                        'CORTES DE ENERGÍA','CAPEX DEL SISTEMA','DEMANDA Y BALANCE']):
            c = ws2.cell(row=i, column=1, value=k)
            c.font = Font(name='Calibri', bold=True, color='F97316', size=10)
            c.fill = HDR_FILL; ws2.merge_cells(f'A{i}:C{i}'); c.border = border
        else:
            data_cell(ws2, i, 1, k)
            c = data_cell(ws2, i, 2, v)
            c.alignment = Alignment(horizontal='right')
            c.font = Font(name='Calibri', bold=True, color='FBBF24', size=9)
            data_cell(ws2, i, 3, u)

    # ── Hoja 3: Ángulos Óptimos ───────────────────────────────────────────────
    ws_ang = wb.create_sheet('3. Ángulos Óptimos')
    ws_ang.sheet_view.showGridLines = False
    ws_ang.column_dimensions['A'].width = 40
    ws_ang.column_dimensions['B'].width = 26
    t_ang = ws_ang.cell(row=1, column=1,
        value='Ángulos Óptimos de Instalación — Análisis Astronómico Solar')
    t_ang.font = TITLE_FONT; ws_ang.merge_cells('A1:B1')

    oa = _cache.get('optimal_angles') or {}
    if not oa and 'solar' in _cache:
        sv = _cache['solar']['stats']
        lat_oa, lon_oa = sv.get('lat', 25.67), sv.get('lon', -100.31)
        try:
            tilt_oa = _optimal_tilt(lat_oa)
            az_oa   = 180.0 if lat_oa >= 0 else 0.0
            hem_oa  = 'Norte' if lat_oa >= 0 else 'Sur'
            ds = _declination_rad(172); de = _declination_rad(80)
            es = _solar_noon_elevation_rad(lat_oa, ds) / _DEG
            ee = _solar_noon_elevation_rad(lat_oa, de) / _DEG
            irr_oa = _estimate_annual_irrad(lat_oa, tilt_oa)
            oa = {
                'lat': lat_oa, 'lon': lon_oa, 'hemisferio': hem_oa,
                'tilt_optimo_deg': tilt_oa, 'azimuth_optimo_deg': az_oa,
                'azimuth_optimo_desc': f"{az_oa:.0f}° — {'Sur' if lat_oa>=0 else 'Norte'}",
                'elevacion_solsticio_deg': round(es, 2),
                'elevacion_equinoccio_deg': round(ee, 2),
                'angulo_cenital_solsticio_deg': round(90 - es, 2),
                'irradiacion_poa_anual_kWh_m2': irr_oa,
                'regla_practica': f"Tilt ≈ {abs(lat_oa):.1f}° ajustado a {tilt_oa:.1f}°",
                'monthly_table': [
                    {
                        'mes': _MONTHS_ES[m],
                        'declinacion_deg': round(_declination_rad(n) / _DEG, 2),
                        'elevacion_solar_deg': round(_solar_noon_elevation_rad(lat_oa, _declination_rad(n)) / _DEG, 2),
                        'angulo_incidencia_deg': round(_incidence_angle_rad(lat_oa, _declination_rad(n), tilt_oa) / _DEG, 2),
                        'cos_theta': round(max(0, math.cos(_incidence_angle_rad(lat_oa, _declination_rad(n), tilt_oa))), 4),
                    }
                    for m, n in enumerate(_MONTH_DAYS)
                ],
            }
        except Exception:
            pass

    ang_params = [
        ('── UBICACIÓN ──', ''),
        ('Latitud analizada', f"{oa.get('lat', '—')} °"),
        ('Longitud', f"{oa.get('lon', '—')} °"),
        ('Hemisferio', oa.get('hemisferio', '—')),
        ('', ''),
        ('── ÁNGULOS ÓPTIMOS ──', ''),
        ('Tilt óptimo anual', f"{oa.get('tilt_optimo_deg', '—')} °"),
        ('Azimut óptimo', oa.get('azimuth_optimo_desc', '—')),
        ('Regla práctica', oa.get('regla_practica', '—')),
        ('', ''),
        ('── ÁNGULOS SOLARES CLAVE ──', ''),
        ('Elevación solar (21 Jun, mediodía)', f"{oa.get('elevacion_solsticio_deg', '—')} °"),
        ('Elevación solar (21 Mar, mediodía)', f"{oa.get('elevacion_equinoccio_deg', '—')} °"),
        ('Ángulo cenital solar (21 Jun)', f"{oa.get('angulo_cenital_solsticio_deg', '—')} °"),
        ('', ''),
        ('── IRRADIACIÓN ESTIMADA ──', ''),
        ('Irradiación POA anual (tilt óptimo)', f"{oa.get('irradiacion_poa_anual_kWh_m2', '—')} kWh/m²·año"),
        ('Modelo utilizado', 'Jensen simplificado — τ_b = 0.7^(AM^0.678)'),
        ('Componente difusa', 'Modelo isotrópico de Hottel-Woertz'),
        ('Albedo suelo', '0.20 (suelo natural)'),
    ]

    for i, (k, v) in enumerate(ang_params, start=3):
        ck = ws_ang.cell(row=i, column=1, value=k)
        cv = ws_ang.cell(row=i, column=2, value=v)
        if k.startswith('──'):
            ck.font = Font(name='Calibri', bold=True, color='10B981', size=10)
            ck.fill = PatternFill("solid", fgColor="0a1a14"); cv.value = ''
        else:
            ck.font = DATA_FONT; cv.font = Font(name='Calibri', bold=True, color='6EE7B7', size=9)
        for c in [ck, cv]:
            c.border = border; c.alignment = Alignment(vertical='center')

    monthly_tbl = oa.get('monthly_table', [])
    if monthly_tbl:
        row_off = len(ang_params) + 4
        t_m = ws_ang.cell(row=row_off, column=1,
            value='Perfil Mensual — Ángulos Solares al Mediodía Solar')
        t_m.font = Font(name='Calibri', bold=True, color='10B981', size=11)
        ws_ang.merge_cells(f'A{row_off}:F{row_off}')
        ang_hdrs = ['Mes','Día Repr.','Declinación δ [°]',
                    'Elevación Solar α [°]','Ángulo Incidencia θ [°]','cos(θ)']
        ang_cols = ['A','B','C','D','E','F']
        col_widths_ang = [16, 12, 20, 22, 24, 12]
        for ci, (h, col, w) in enumerate(zip(ang_hdrs, ang_cols, col_widths_ang), 1):
            hdr(ws_ang, row_off + 1, ci, h)
            ws_ang.column_dimensions[col].width = w
        best_cos_row = max(monthly_tbl, key=lambda x: x.get('cos_theta', 0)) if monthly_tbl else {}
        for ri, row_d in enumerate(monthly_tbl):
            r = row_off + 2 + ri
            vals = [row_d.get('mes',''), row_d.get('dia_repr',''),
                    row_d.get('declinacion_deg',''), row_d.get('elevacion_solar_deg',''),
                    row_d.get('angulo_incidencia_deg',''), row_d.get('cos_theta','')]
            is_best = row_d.get('mes') == best_cos_row.get('mes')
            for ci, val in enumerate(vals, 1):
                c = data_cell(ws_ang, r, ci, val,
                              fmt='0.00' if ci > 1 else None,
                              fill=PatternFill("solid", fgColor="0a1a14"))
                if is_best:
                    c.font = Font(name='Calibri', bold=True, color='FBBF24', size=9)

    # ── Hoja 4: Resumen Mensual ────────────────────────────────────────────────
    ws3 = wb.create_sheet('4. Resumen Mensual')
    ws3.sheet_view.showGridLines = False
    MONTHS_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                 'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    t3 = ws3.cell(row=1, column=1, value='Resumen Mensual — Generación, Temperatura y Balance')
    t3.font = TITLE_FONT; ws3.merge_cells('A1:I1')
    headers3 = ['Mes','Irrad. POA Avg\n[W/m²]','Irrad. POA Máx\n[W/m²]',
                'Gen. Neta\n[kWh/mes]','Bat. Cargada\n[kWh/mes]',
                'Bat. Descargada\n[kWh/mes]','Demanda\n[kWh/mes]',
                'Cobertura\n[%]','Balance\n[kWh/mes]']
    for ci, h in enumerate(headers3, 1):
        hdr(ws3, 2, ci, h); ws3.column_dimensions[get_column_letter(ci)].width = 16

    dem_mon = [0] * 12
    if demand:
        dem_arr = np.array(demand['demand_kW']); idx_ = 0
        for mi_, nd in enumerate(DAYS_IN_MONTH):
            seg = dem_arr[idx_:idx_ + nd * 96]
            dem_mon[mi_] = float(np.sum(seg) * 0.25); idx_ += nd * 96

    bat_ch  = solar.get('monthly_bat_charge',   [0]*12)
    bat_dis = solar.get('monthly_bat_discharge', [0]*12)
    for mi, mes in enumerate(MONTHS_ES):
        r   = mi + 3
        gen = solar['monthly_gen_kWh'][mi]
        dem = dem_mon[mi]
        cob = min(gen / dem * 100, 100) if dem > 0 else 0
        row_d = [mes,
                 round(solar['monthly_gtot_avg'][mi], 1),
                 round(solar['monthly_gtot_max'][mi], 1),
                 round(gen, 0), round(bat_ch[mi], 2), round(bat_dis[mi], 2),
                 round(dem, 0), round(cob, 1), round(gen - dem, 0)]
        fills = [None, COL_FILLS['irrad'], COL_FILLS['irrad'],
                 COL_FILLS['gen'], COL_FILLS['gen'], COL_FILLS['gen'],
                 COL_FILLS['dem'], COL_FILLS['bal'], COL_FILLS['bal']]
        fmts  = [None,'#,##0.0','#,##0.0','#,##0','#,##0.00','#,##0.00',
                 '#,##0','0.0"%"','#,##0']
        for ci, (val, fill, fmt) in enumerate(zip(row_d, fills, fmts), 1):
            c = data_cell(ws3, r, ci, val, fmt=fmt, fill=fill)
            if ci == 1: c.font = Font(name='Calibri', bold=True, color='F1F5F9', size=9)

    # ── Hoja 5: Datos Anuales ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('5. Datos Anuales (15-min)')
    ws4.sheet_view.showGridLines = False
    headers4 = ['Fecha-Hora','Irrad. POA\n[W/m²]','Gen. Neta\n[kW]',
                'Gen. Bruta\n[kW]','T. Celda\n[°C]','Demanda\n[kW]','Balance\n[kW]']
    col_widths4 = [18, 15, 13, 13, 11, 13, 13]
    for ci, (h, w) in enumerate(zip(headers4, col_widths4), 1):
        hdr(ws4, 1, ci, h); ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = 'A2'
    fills4 = [COL_FILLS['fecha'], COL_FILLS['irrad'],
              COL_FILLS['gen'],   COL_FILLS['gen'],
              COL_FILLS['temp'],  COL_FILLS['dem'], COL_FILLS['bal']]
    fmts4  = [None,'#,##0.0','#,##0.00','#,##0.00','#,##0.0','#,##0.00','#,##0.00']
    dem_kw_arr = dem_kw if demand else [0.0] * len(hours)
    T_arr  = solar.get('T_cell_arr', [0.0] * len(hours))
    for i, (f, g, p, pg, tc, dm) in enumerate(
            zip(fechas, Gtot, P_kw, P_gross, T_arr, dem_kw_arr), start=2):
        bal = p - dm
        row_v = [f, round(g,2), round(p,4), round(pg,4), round(tc,1), round(dm,4), round(bal,4)]
        for ci, (val, fill, fmt) in enumerate(zip(row_v, fills4, fmts4), 1):
            data_cell(ws4, i, ci, val, fmt=fmt, fill=fill)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='Motor_Solar_FV_v2.1_Resultados.xlsx')


@app.route('/api/download', methods=['GET'])
def api_download_csv():
    return api_download_excel()


@app.route('/api/cfe_gdmto', methods=['GET'])
def api_cfe_gdmto():
    """Obtiene tarifas GDMTO en vivo desde app.cfe.mx (SSL-broken, usa verify=False)."""
    FALLBACK_DIVISIONES = [
        {'nombre': 'CDMX / Valle de México (referencia 2025)', 'precio_kwh': 1.699, 'cargo_fijo': 466.83, 'cargo_demanda': 437.87},
        {'nombre': 'Noroeste (referencia 2025)',                'precio_kwh': 1.821, 'cargo_fijo': 466.83, 'cargo_demanda': 450.12},
        {'nombre': 'Norte (referencia 2025)',                   'precio_kwh': 1.756, 'cargo_fijo': 466.83, 'cargo_demanda': 442.55},
    ]
    fecha_hoy = datetime.date.today().isoformat()

    URL = "https://app.cfe.mx/Aplicaciones/CCFE/Tarifas/TarifasCRENegocio/Tarifas/GranDemandaMTO.aspx"

    try:
        r = requests.get(
            URL, verify=False, timeout=12,
            headers={'User-Agent': 'Mozilla/5.0 (compatible; SolarCalc/2.2)'}
        )
        r.raise_for_status()
        r.encoding = r.apparent_encoding or 'utf-8'
        soup = BeautifulSoup(r.text, 'html.parser')

        money_re = re.compile(r'\$?\s*(\d{1,6}(?:,\d{3})*(?:\.\d+)?)')
        divisiones = []

        for table in soup.find_all('table'):
            for row in table.find_all('tr'):
                cells = row.find_all(['td', 'th'])
                if len(cells) < 3:
                    continue
                row_text = ' '.join(c.get_text(strip=True) for c in cells)
                nums_raw = money_re.findall(row_text)
                nums = [float(n.replace(',', '')) for n in nums_raw if float(n.replace(',', '')) > 0.01]
                if len(nums) < 2:
                    continue
                nombre = cells[0].get_text(strip=True)
                if not nombre or nombre.lower() in ('concepto', 'división', 'division', 'region', 'región'):
                    continue
                # Columnas esperadas CFE GDMTO: [0] cargo_fijo | [1] precio_kwh | [2] cargo_dist | [3] cargo_cap
                # Si la página cambia de estructura el fallback entra automáticamente (divisiones queda vacío).
                cargo_fijo    = nums[0] if len(nums) > 0 else FALLBACK_DIVISIONES[0]['cargo_fijo']
                precio_kwh    = nums[1] if len(nums) > 1 else FALLBACK_DIVISIONES[0]['precio_kwh']
                cargo_demanda = (nums[2] + nums[3]) if len(nums) > 3 else (nums[2] if len(nums) > 2 else FALLBACK_DIVISIONES[0]['cargo_demanda'])
                divisiones.append({'nombre': nombre, 'precio_kwh': precio_kwh, 'cargo_fijo': cargo_fijo, 'cargo_demanda': cargo_demanda})

        if not divisiones:
            return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback', 'fecha': fecha_hoy})

        return jsonify({'ok': True, 'divisiones': divisiones, 'fuente': URL, 'fecha': fecha_hoy})

    except Exception as e:
        return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback',
                        'error_detail': str(e), 'fecha': fecha_hoy})


@app.route('/api/cfe_gdmto_tarifa', methods=['POST'])
def api_cfe_gdmto_tarifa():
    """Multi-step WebForms POST para tarifas GDMTO según ubicación y mes/año."""
    FALLBACK_DIVISIONES = [
        {'nombre': 'CDMX / Valle de México (respaldo 2025)', 'precio_kwh': 1.699, 'cargo_fijo': 466.83, 'cargo_demanda': 437.87},
        {'nombre': 'Noroeste (respaldo 2025)',                'precio_kwh': 1.821, 'cargo_fijo': 466.83, 'cargo_demanda': 450.12},
        {'nombre': 'Norte (respaldo 2025)',                   'precio_kwh': 1.756, 'cargo_fijo': 466.83, 'cargo_demanda': 442.55},
    ]
    fecha_hoy = datetime.date.today().isoformat()
    URL = 'https://app.cfe.mx/Aplicaciones/CCFE/Tarifas/TarifasCRENegocio/Tarifas/GranDemandaMTO.aspx'
    HDR = {'User-Agent': 'Mozilla/5.0 (compatible; SolarCalc/2.3)'}

    body = request.get_json(silent=True) or {}
    lat  = float(body.get('lat',  25.67))
    lon  = float(body.get('lon', -100.31))
    anio = int(body.get('anio', datetime.date.today().year))
    mes  = int(body.get('mes',  datetime.date.today().month))

    # ── Reverse geocoding via Google Maps ──────────────────────────────
    GMAPS_KEY = 'AIzaSyDIO9AKyM4TeZJ2O2uLbgPETJapKZLo_d4'
    estado_name = None
    municipio_name = None
    try:
        geo_r = requests.get(
            f'https://maps.googleapis.com/maps/api/geocode/json'
            f'?latlng={lat},{lon}&key={GMAPS_KEY}&language=es',
            timeout=5
        )
        geo_d = geo_r.json()
        if geo_d.get('status') == 'OK' and geo_d.get('results'):
            for comp in geo_d['results'][0].get('address_components', []):
                types = comp.get('types', [])
                if 'administrative_area_level_1' in types and estado_name is None:
                    estado_name = comp['long_name'].upper()
                if ('locality' in types or 'administrative_area_level_2' in types) and municipio_name is None:
                    municipio_name = comp['long_name'].upper()
    except Exception:
        pass

    def _norm(s):
        return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii').upper()

    def best_match(target, options):
        if target:
            nt = _norm(target)
            for val, text in options:
                if val and val != '0' and (_norm(text) and (nt in _norm(text) or _norm(text) in nt)):
                    return val
        for val, text in options:
            if val and val != '0':
                return val
        return None

    def extract_hidden(soup, name):
        el = soup.find('input', {'name': name})
        return el['value'] if el else ''

    def extract_options(soup, select_name):
        sel = soup.find('select', {'name': select_name})
        if not sel:
            return []
        return [(o.get('value', ''), o.get_text(strip=True)) for o in sel.find_all('option')]

    try:
        sess = requests.Session()

        # ── Step 1: GET page ────────────────────────────────────────────
        r1 = sess.get(URL, verify=False, timeout=12, headers=HDR)
        r1.raise_for_status()
        r1.encoding = r1.apparent_encoding or 'utf-8'
        s1 = BeautifulSoup(r1.text, 'html.parser')

        vs  = extract_hidden(s1, '__VIEWSTATE')
        vsg = extract_hidden(s1, '__VIEWSTATEGENERATOR')
        ev  = extract_hidden(s1, '__EVENTVALIDATION')

        estado_options = extract_options(s1, 'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddEstado')
        estado_val = best_match(estado_name, estado_options)
        if not estado_val:
            return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback', 'fecha': fecha_hoy})

        # ── Step 2: POST estado → get municipios ─────────────────────────
        base_fields = {
            '__VIEWSTATE':          vs,
            '__VIEWSTATEGENERATOR': vsg,
            '__EVENTVALIDATION':    ev,
            '__EVENTARGUMENT':      '',
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddEstado':    estado_val,
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddMunicipio': '0',
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddDivision':  '0',
            'ctl00$ContentPlaceHolder1$Fecha$ddAnio':          str(anio),
            'ctl00$ContentPlaceHolder1$Fecha2$ddMes':          str(mes),
            'ctl00$ContentPlaceHolder1$hdAnio':                '',
            'ctl00$ContentPlaceHolder1$hdMes':                 '',
        }
        r2 = sess.post(URL, data={**base_fields,
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddEstado'},
            verify=False, timeout=15, headers=HDR)
        s2 = BeautifulSoup(r2.text, 'html.parser')
        vs = extract_hidden(s2, '__VIEWSTATE') or vs
        ev = extract_hidden(s2, '__EVENTVALIDATION') or ev

        municipio_options = extract_options(s2, 'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddMunicipio')
        municipio_val = best_match(municipio_name, municipio_options)
        if not municipio_val:
            return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback', 'fecha': fecha_hoy})

        # ── Step 3: POST municipio → get divisions ───────────────────────
        r3 = sess.post(URL, data={**base_fields,
            '__VIEWSTATE':       vs,
            '__EVENTVALIDATION': ev,
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddMunicipio',
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddMunicipio': municipio_val},
            verify=False, timeout=15, headers=HDR)
        s3 = BeautifulSoup(r3.text, 'html.parser')
        vs = extract_hidden(s3, '__VIEWSTATE') or vs
        ev = extract_hidden(s3, '__EVENTVALIDATION') or ev

        division_options = extract_options(s3, 'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddDivision')
        division_val = best_match(None, division_options)
        if not division_val:
            return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback', 'fecha': fecha_hoy})
        division_nombre = next((t for v, t in division_options if v == division_val), division_val)

        # ── Step 4: POST final → get tariff table ────────────────────────
        submit_btn = s3.find('input', {'type': 'submit'})
        btn_name  = submit_btn['name']  if submit_btn and submit_btn.get('name')  else 'ctl00$ContentPlaceHolder1$btnConsultar'
        btn_value = submit_btn['value'] if submit_btn and submit_btn.get('value') else 'Consultar'

        r4 = sess.post(URL, data={
            '__VIEWSTATE':          vs,
            '__VIEWSTATEGENERATOR': vsg,
            '__EVENTVALIDATION':    ev,
            '__EVENTTARGET':        '',
            '__EVENTARGUMENT':      '',
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddEstado':    estado_val,
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddMunicipio': municipio_val,
            'ctl00$ContentPlaceHolder1$EdoMpoDiv$ddDivision':  division_val,
            'ctl00$ContentPlaceHolder1$Fecha$ddAnio':          str(anio),
            'ctl00$ContentPlaceHolder1$Fecha2$ddMes':          str(mes),
            'ctl00$ContentPlaceHolder1$hdAnio':                '',
            'ctl00$ContentPlaceHolder1$hdMes':                 '',
            btn_name:                                          btn_value,
        }, verify=False, timeout=20, headers=HDR)
        s4 = BeautifulSoup(r4.text, 'html.parser')

        # ── Parse tariff table ───────────────────────────────────────────
        # CFE GDMTO has 4 charges: Fijo ($/mes), Variable/Energía ($/kWh),
        # Distribución ($/kW), Capacidad ($/kW).  They may appear in the same
        # row or in separate rows — so we search for each keyword and take the
        # first number that follows it in the text (within 80 chars).
        full_text = ' '.join(s4.stripped_strings).upper()

        def _val_after(text, keyword, min_val=0.0, max_val=1e9, window=80):
            idx = text.find(keyword)
            if idx == -1:
                return None
            segment = text[idx + len(keyword): idx + len(keyword) + window]
            m = re.search(r'(\d[\d,]*\.?\d*)', segment)
            if not m:
                return None
            v = float(m.group(1).replace(',', ''))
            return v if min_val < v <= max_val else None

        fijo_val = _val_after(full_text, 'FIJO',      min_val=50)
        kwh_val  = _val_after(full_text, 'VARIABLE',  max_val=20) or \
                   _val_after(full_text, 'ENERG',     max_val=20)
        dist_val = _val_after(full_text, 'DISTRIBUCI', min_val=1)
        cap_val  = _val_after(full_text, 'CAPACIDAD',  min_val=1)

        cargo_demanda_total = (dist_val or 0) + (cap_val or 0)
        divisiones = []
        if kwh_val is not None:
            divisiones = [{
                'nombre':       division_nombre,
                'precio_kwh':   kwh_val,
                'cargo_fijo':   fijo_val if fijo_val is not None else FALLBACK_DIVISIONES[0]['cargo_fijo'],
                'cargo_demanda': cargo_demanda_total if cargo_demanda_total > 1 else FALLBACK_DIVISIONES[0]['cargo_demanda'],
            }]

        if not divisiones:
            return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback', 'fecha': fecha_hoy})

        return jsonify({'ok': True, 'divisiones': divisiones, 'fuente': URL,
                        'fecha': fecha_hoy, 'division': division_nombre})

    except Exception as e:
        return jsonify({'ok': True, 'divisiones': FALLBACK_DIVISIONES, 'fuente': 'fallback',
                        'error_detail': str(e), 'fecha': fecha_hoy})


if __name__ == '__main__':
    print("=" * 60)
    print("  Motor Solar Fotovoltaico v2.1 — Servidor Flask")
    print("  Abrir en navegador: http://localhost:8000")
    print("=" * 60)
    app.run(debug=True, port=8000, use_reloader=False)